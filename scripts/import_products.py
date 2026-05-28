# scripts/import_products.py

# Скрипт для импорта средств из Excel-файла в базу данных.

import json
import sys
from pathlib import Path

# Добавляем корень проекта в sys.path, чтобы импортировать core и apps
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from dotenv import load_dotenv

load_dotenv()

from sqlalchemy import select

from apps.products.models import Product, ProductFillStatus
from core.database import sync_session

EXCEL_PATH = Path(__file__).parent.parent / "table" / "Таблица средств.xlsx"
SHEET_NAME = "Средства"

COLUMN_MAP = {
    "Ссылка в Золотом Яблоке": "link_ga",
    "Тип продукта": "product_type",
    "Тип продукта подробно": "product_type_detailed",
    "Название": "name",
    "Артикул в Золотом Яблоке": "article_ga",
    "Назначение": "purpose",
    "Тип волос": "hair_type",
    "Тип кожи": "skin_type",
    "Область применения": "application_area",
    "Состав": "ingredients",
    "Элементы состава списком": "ingredients_list",
    "Количество элементов состава": "ingredients_count",
    "Для кого": "target_audience",
    "Мера (объём/количество)": "measure_type",
    "Количество меры (число)": "measure_value",
    "Юниты меры (мл/шт)": "measure_unit",
    "Стоимость руб": "price_rub",
    "Описание": "description",
    "Применение": "usage",
    "Бренд": "brand",
    "Страна бренда": "brand_country",
    "Описание бренда": "brand_description",
    "Дополнительная информация": "additional_info",
    "Ссылка на изображение": "image_link",
    "Характеристики": "characteristics",
    "Заполнено": "fill_status",
    "Ключ на изображение на S3": "image_key",
}


def _parse_row(raw_data: dict) -> dict:
    """
    Преобразует сырую строку из Excel в словарь для модели Product.
    """
    product_data = {}

    for column_name, field_name in COLUMN_MAP.items():
        value = raw_data.get(column_name)

        if field_name == "ingredients_list":
            if isinstance(value, str):
                try:
                    value = json.loads(value)
                except json.JSONDecodeError as e:
                    print(f"⚠️  Не удалось распарсить ingredients_list: {e}")
                    value = []
            else:
                value = []

        elif field_name == "fill_status":
            is_filled = str(value).strip().lower() in ("1", "true", "да")
            value = ProductFillStatus.DONE if is_filled else ProductFillStatus.EMPTY

        product_data[field_name] = value

    return product_data


def import_products() -> None:
    """
    Читает Excel-файл и выполняет upsert продуктов в БД по полю link_ga.
    Каждая строка — отдельный savepoint: ошибка в одной строке не откатывает остальные.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb[SHEET_NAME]

    header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
    headers = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in header_cells
        if cell.value is not None
    ]

    error_count = 0

    with sync_session() as session:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw_data = dict(zip(headers, row))

            # Пропускаем пустые строки
            if not raw_data.get("Ссылка в Золотом Яблоке"):
                continue

            product_data = _parse_row(raw_data)

            try:
                with session.begin_nested():  # savepoint — ошибка не откатит всё
                    existing = session.execute(
                        select(Product).where(
                            Product.link_ga == product_data["link_ga"]
                        )
                    ).scalar_one_or_none()

                    if existing:
                        for key, val in product_data.items():
                            setattr(existing, key, val)
                    else:
                        session.add(Product(**product_data))

            except Exception as e:
                error_count += 1
                print(f"❌ Ошибка при обработке: {product_data.get('name')!r}")
                print(f"   {e}")

        session.commit()

    print(f"\n✅ Импорт завершён. Ошибок: {error_count}")


if __name__ == "__main__":
    import_products()
