# scripts/upload_images_to_s3.py

# Скрипт для автоматической загрузки изображений, указанных в Excel-файле,
# в облачное хранилище S3 с сохранением обратной ссылки в Excel.
# Запускается из entrypoint.sh при старте контейнера.
# Для ручного запуска из корня проекта:
#   python scripts/upload_images_to_s3.py

import asyncio
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from dotenv import load_dotenv

load_dotenv()

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from infrastructure.s3.service import S3Service

# Путь до Excel
EXCEL_PATH = Path("table/Таблица средств.xlsx")
SHEET_NAME = "Средства"

# Названия колонок
IMAGE_PATH_COLUMN = "Ссылка на изображение в базе"
S3_KEY_COLUMN = "Ключ на изображение на S3"

# Путь до корня проекта
PROJECT_ROOT = Path(__file__).resolve().parent.parent


async def upload_single_image(
    row: tuple[Cell, ...],
    row_index: int,
    image_col_idx: int,
    s3_key_col_idx: int,
    s3: S3Service,
) -> bool:
    """
    :param row: Строка Excel-файла (объект openpyxl).
    :param row_index: Номер текущей строки (для логов).
    :param image_col_idx: Номер колонки с путём до изображения.
    :param s3_key_col_idx: Номер колонки, куда будет записан ключ S3.
    :param s3: Экземпляр сервиса работы с S3.
    :return: True, если что-то изменилось (загрузка или новый ключ в ячейке),
        иначе False — чтобы не пересохранять Excel зря.
    """
    # Получаем значение из колонки с путём до изображения
    image_path = row[image_col_idx - 1].value

    if not image_path:
        return False

    # Ячейка, в которую будем писать ключ S3
    s3_key_cell = row[s3_key_col_idx - 1]

    # Формируем абсолютный путь к изображению на диске
    local_path = PROJECT_ROOT / image_path

    # Собираем ключ для хранения на S3 (только имя файла)
    s3_key = f"images/{Path(image_path).name}"

    # Проверяем наличие на S3 HEAD-запросом (без скачивания тела файла)
    response = await s3.file_exists(s3_key)
    if response["status_code"] == 200:
        print(f"🔁 [{row_index}] Уже на S3, пропускаем: {s3_key}")
        # Проставляем ключ, только если он ещё не записан — иначе менять нечего
        if s3_key_cell.value != s3_key:
            s3_key_cell.value = s3_key
            return True
        return False

    # Пробуем открыть локальный файл
    try:
        with local_path.open("rb") as f:
            file_data = f.read()
    except FileNotFoundError:
        print(f"❌ [{row_index}] Файл не найден: {local_path}")
        return False

    # Получаем расширение файла (например, "jpg")
    extension = local_path.suffix.lstrip(".")

    # Загружаем файл на S3
    result = await s3.upload_file(file_data, s3_key, extension)

    if result["error"]:
        print(f"❌ [{row_index}] Ошибка при загрузке: {result['error']}")
        return False

    print(f"✅ [{row_index}] Загружено: {s3_key}")
    s3_key_cell.value = s3_key
    return True


async def upload_images_from_excel_to_s3() -> None:
    """
    Автоматическая загрузка изображений, указанных в Excel-файле,
    в облачное хранилище S3.
    """
    s3 = S3Service()

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    headers = [cell.value for cell in ws[1]]
    image_col_idx = headers.index(IMAGE_PATH_COLUMN) + 1
    s3_key_col_idx = headers.index(S3_KEY_COLUMN) + 1

    changed = False
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row_index = row[0].row
        row_changed = await upload_single_image(
            row=row,
            row_index=row_index,
            image_col_idx=image_col_idx,
            s3_key_col_idx=s3_key_col_idx,
            s3=s3,
        )
        changed = changed or row_changed

    # Сохраняем Excel только если реально что-то поменялось —
    # иначе при каждом старте контейнера файл переписывался бы зря.
    if changed:
        wb.save(EXCEL_PATH)
        print("💾 Excel обновлён")
    else:
        print("💾 Excel без изменений — сохранение пропущено")


if __name__ == "__main__":
    asyncio.run(upload_images_from_excel_to_s3())
